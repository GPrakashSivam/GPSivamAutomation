import pandas as pd
import openpyxl
import requests
import gspread
import re
from bs4 import BeautifulSoup
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils.dataframe import dataframe_to_rows
import commonUtility as common

#Program to download data file from web, process the data and get Camarilla values
#and finally store those values in local XL file and google sheets.

def downloadWebData():
    print("Getting data from web...")
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/56.0.2924.76 Safari/537.36'}
    response = requests.get(common.webURL, headers=headers)
    soup = BeautifulSoup(response.text,'html.parser')
    # Bhavcopy trade data stored in vBC variable in the script tag of page content
    regex = r"var vBC=(.*?);"
    scripts = soup.find_all('script')
    for script in scripts:	
       match = re.search(regex, str(script.string)) 
       if match != None: 
          df = pd.read_json(match.group(1))
          df = df[df['InstrumentName']=='FUTCOM'] #filter only FUTCOM symbols
          df = df.iloc[:, 1:8] #only get columns 1 to 8 from dataframe
          break
    return df
#EOF downloadWebData

def processInputData(data):
    print("Processing data...")

    #Strip of leading whitespaces in the column
    data["Symbol"] = data["Symbol"].str.strip()
    symbols = common.symbols.split(",")
    # populate dataframe with first record
    finalData = filterData(data,symbols[0])

    # append dataframe from second record onwards
    for i in range(1,len(symbols)):
        finalData = finalData.append(filterData(data,symbols[i]))

    # reseting the default index
    finalData.set_index(['Symbol'], inplace = True)
    return finalData
#EOF-processInputData

def filterData(data,options):
    options = options.split("-")
    symbol = options[0]
    rownum = options[1]
    #selecting rows based on condition and then select top row from the rows
    if(rownum == "1"):
        data = (data[data['Symbol']==symbol]).head(1)
    else:
        data = ((data[data['Symbol']==symbol]).head(2)).tail(1)
    return data
#EOF filterData

def storeDataToXL(data):
    workbook = openpyxl.load_workbook(common.outputXLFilename)
    worksheet = workbook['Sheet2']

    #Saving OHLC values to Excel in Sheet2
    rows = dataframe_to_rows(data, index=False, header=False)
    for r_idx, row in enumerate(rows, 2):
        for c_idx, value in enumerate(row, 2):
             worksheet.cell(row=r_idx, column=c_idx, value=value)

    #Calculating High Values
    data['H6'] = (data['High']/data['Low'])*data['Close']
    data['H4'] = (0.55*(data['High'] - data['Low'])) + data['Close']
    data['H5'] = (data['H6'] + data['H4'])/2
    data['H3'] = (0.275*(data['High'] - data['Low'])) + data['Close']
    data['H2'] = (0.183*(data['High'] - data['Low'])) + data['Close']
    data['H1'] = (0.0916*(data['High'] - data['Low'])) + data['Close']
    #Calculating Low Values
    data['L1'] = data['Close'] - (0.0916*(data['High'] - data['Low']))
    data['L2'] = data['Close'] - (0.183*(data['High'] - data['Low']))
    data['L3'] = data['Close'] - (0.275*(data['High'] - data['Low']))
    data['L4'] = data['Close'] - (0.55*(data['High'] - data['Low']))
    data['L6'] = data['Close'] - (data['H6']- data['Close'])
    data['L5'] = (data['L4'] + data['L6'])/2

    data = round(data,2)
    data = data[['H6','H5','H4','H3','H2','H1','L1','L2','L3','L4','L5','L6']]

    #Saving Camarilla values to Excel in Sheet1
    worksheet = workbook['Sheet1']
    rows = dataframe_to_rows(data, index=False, header=False)
    for row_idx, row in enumerate(rows, 3):
        for col_idx, value in enumerate(row, 3):
            worksheet.cell(row=col_idx, column=row_idx, value=value)

    workbook.save(common.outputXLFilename)
    print('Final Data stored to Excel sucessfully...')
    return data
#EOF storeDataToXL

def saveDataToGoogleSheets(data):
    scope = ["https://spreadsheets.google.com/feeds", 'https://www.googleapis.com/auth/spreadsheets', 
		"https://www.googleapis.com/auth/drive.file", "https://www.googleapis.com/auth/drive"] 

    # Assign credentials ann path of style sheet 
    creds = ServiceAccountCredentials.from_json_keyfile_name(("GprakashSivam.json"), scope) 
    client = gspread.authorize(creds) 
    sheets = client.open(common.googleSheetName)
    worksheet = sheets.get_worksheet(0)
    dataTranposed = data.transpose().fillna(0)
    worksheet.update('CamarillaRange', dataTranposed.values.tolist())
    print("Data moved to Google sheets...")
#EOF saveDataToGoogleSheets

if __name__=="__main__": 
    try:
        data = downloadWebData()
        data = processInputData(data)
        data = storeDataToXL(data)
        saveDataToGoogleSheets(data)
    except Exception as e:
        print("Error: ",e)
    finally:
        print("Exiting program...")
